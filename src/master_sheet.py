"""Utilities to load and normalize the Master Excel into in-memory dicts.
This module provides a stable shape so the rest of the codebase can accept either the real Master Excel or the placeholder.
"""
from openpyxl import load_workbook
from pathlib import Path
from typing import List, Dict, Any


def _row_as_dict(header, row):
    d = {}
    for h, cell in zip(header, row):
        val = cell.value
        # normalize ints/floats where possible
        if isinstance(val, (int, float)):
            d[h] = val
        else:
            # try convert
            try:
                if val is None:
                    d[h] = None
                else:
                    if str(val).replace('.', '', 1).isdigit():
                        if '.' in str(val):
                            d[h] = float(val)
                        else:
                            d[h] = int(val)
                    else:
                        d[h] = val
            except Exception:
                d[h] = val
    return d


def load_master_excel(path: Path) -> Dict[str, List[Dict[str, Any]]]:
    wb = load_workbook(filename=str(path), data_only=True)
    sheets = {}
    for name in wb.sheetnames:
        ws = wb[name]
        rows = list(ws.rows)
        if not rows:
            sheets[name] = []
            continue
        header = [c.value for c in rows[0]]
        data = []
        for r in rows[1:]:
            data.append(_row_as_dict(header, r))
        sheets[name] = data
    return sheets


def validate_master_data(sheets: Dict[str, List[Dict[str, Any]]]) -> Dict[str, Any]:
    """Run simple validation checks across common tabs and return a report.
    This is intentionally lightweight but extendable to full formula-by-formula checks.
    """
    mismatches = []
    summary = {"hardware_rows": 0, "structural_rows": 0, "labor_rows": 0}

    hardware = sheets.get("Hardware", [])
    structural = {r.get("screen_id"): r for r in sheets.get("Structural", [])}
    labor = {r.get("screen_id"): r for r in sheets.get("Labor", [])}

    for row in hardware:
        summary["hardware_rows"] += 1
        sid = row.get("screen_id")
        sqft = row.get("sqft") or (row.get("width_ft", 0) * row.get("height_ft", 0))
        # Support alternative header name used in some master sheets
        base = row.get("base_rate") or row.get("base_rate_per_sqft") or 0
        hw_cost = row.get("hardware_cost") or 0
        expected_hw = float(sqft) * float(base)
        if abs(expected_hw - float(hw_cost)) > 0.01:
            mismatches.append({"screen_id": sid, "field": "hardware_cost", "expected": expected_hw, "found": hw_cost})

        # Structural check
        s = structural.get(sid)
        if s:
            summary["structural_rows"] += 1
            expected_struct = expected_hw * (s.get("structural_pct") or 0)
            if abs(expected_struct - float(s.get("structural_cost", 0))) > 0.01:
                mismatches.append({"screen_id": sid, "field": "structural_cost", "expected": expected_struct, "found": s.get("structural_cost")})

        # Labor check
        l = labor.get(sid)
        if l:
            summary["labor_rows"] += 1
            expected_labor = (expected_hw + (s.get("structural_cost") if s else 0)) * (l.get("labor_pct") or 0)
            if abs(expected_labor - float(l.get("labor_cost", 0))) > 0.01:
                mismatches.append({"screen_id": sid, "field": "labor_cost", "expected": expected_labor, "found": l.get("labor_cost")})

    return {"mismatches": mismatches, "summary": summary}

if __name__ == '__main__':
    p = Path('samples/master_excel_dummy.xlsx')
    if not p.exists():
        print('No master sheet found. Run scripts/generate_dummy_master_excel.py to create a placeholder.')
    else:
        d = load_master_excel(p)
        for k, v in d.items():
            print(f"{k}: {len(v)} rows")
