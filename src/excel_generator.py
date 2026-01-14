import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


class ExcelGenerator:
    def __init__(self):
        pass

    def update_expert_estimator(
        self, project_data, output_path="anc_internal_estimation.xlsx"
    ):
        """
        Generates a multi-tab Excel file matching ANC's internal workflow.
        Based on Natalia's requirements: separate tabs for LED, Structural, Labor, Electrical, etc.
        """
        wb = openpyxl.Workbook()

        # Remove default sheet if it exists
        if wb.active:
            wb.remove(wb.active)

        # 1. Executive Summary Tab - Client-facing overview
        ws_summary = wb.create_sheet(title="Executive Summary")
        self._generate_executive_summary_tab(ws_summary, project_data)

        # 2. LED Hardware Tab - Display specifications and costs
        ws_led = wb.create_sheet(title="LED Hardware")
        self._generate_led_hardware_tab(ws_led, project_data)

        # 3. Structural Requirements Tab - Steel, mounting, engineering
        ws_structural = wb.create_sheet(title="Structural Requirements")
        self._generate_structural_tab(ws_structural, project_data)

        # 4. Labor Analysis Tab - Union rates, crew sizes, installation complexity
        ws_labor = wb.create_sheet(title="Labor Analysis")
        self._generate_labor_tab(ws_labor, project_data)

        # 5. Electrical Systems Tab - Power, data, PDUs
        ws_electrical = wb.create_sheet(title="Electrical Systems")
        self._generate_electrical_tab(ws_electrical, project_data)

        # 6. Installation Assessment Tab - Site conditions, access, complexity
        ws_installation = wb.create_sheet(title="Installation Assessment")
        self._generate_installation_tab(ws_installation, project_data)

        # 7. Professional Services Tab - PM, engineering, permits
        ws_services = wb.create_sheet(title="Professional Services")
        self._generate_professional_services_tab(ws_services, project_data)

        # 8. Cost Summary Tab - Detailed breakdown by screen
        for idx, screen_data in enumerate(project_data):
            sheet_name = f"Screen {idx + 1} Details"
            ws_detail = wb.create_sheet(title=sheet_name)
            self._generate_screen_detail_tab(ws_detail, screen_data, idx + 1)

        wb.save(output_path)
        print(f"Excel generated at: {output_path}")

    def _generate_executive_summary_tab(self, ws, project_data):
        """Executive summary for management review"""
        # Header styling
        header_font = Font(bold=True, color="FFFFFF", size=14)
        header_fill = PatternFill(
            start_color="003D82", end_color="003D82", fill_type="solid"
        )

        # Title
        ws.cell(
            row=1, column=1, value="ANC SPORTS ENTERPRISES - PROJECT SUMMARY"
        ).font = Font(bold=True, size=16)
        ws.merge_cells("A1:F1")

        # Project overview
        headers = [
            "Screen #",
            "Product Type",
            "Dimensions",
            "Pixel Pitch",
            "Total Cost",
            "Annual Service",
        ]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[get_column_letter(col)].width = 20

        # Data rows
        total_cost = 0
        for idx, item in enumerate(project_data):
            row = idx + 4
            inp = item["inputs"]
            summary = item["summary"]

            cost = summary.get("final_sell_price", 0)
            service = summary.get("annual_service", 0)
            total_cost += cost

            ws.cell(row=row, column=1, value=idx + 1)
            ws.cell(row=row, column=2, value=inp.product_class)
            ws.cell(row=row, column=3, value=f"{inp.width_ft}' × {inp.height_ft}'")
            ws.cell(row=row, column=4, value=f"{inp.pixel_pitch}mm")
            ws.cell(row=row, column=5, value=cost).number_format = "$#,##0"
            ws.cell(row=row, column=6, value=service).number_format = "$#,##0"

        # Totals
        total_row = len(project_data) + 5
        ws.cell(row=total_row, column=1, value="PROJECT TOTAL").font = Font(bold=True)
        ws.cell(row=total_row, column=5, value=total_cost).number_format = "$#,##0"
        ws.cell(row=total_row, column=5).font = Font(bold=True, size=14)

    def _generate_led_hardware_tab(self, ws, project_data):
        """LED display specifications and hardware costs"""
        ws.cell(row=1, column=1, value="LED HARDWARE SPECIFICATIONS").font = Font(
            bold=True, size=14
        )

        headers = [
            "Screen",
            "Product",
            "Pixel Pitch",
            "Dimensions",
            "Resolution",
            "Sq Ft",
            "Cost/Sq Ft",
            "Hardware Cost",
        ]
        self._set_headers(ws, headers, 3)

        total_hardware = 0
        for idx, item in enumerate(project_data):
            row = idx + 4
            inp = item["inputs"]
            details = item["details"]

            hardware_cost = details.get("hardware", {}).get("raw_cost", 0)
            total_hardware += hardware_cost

            ws.cell(row=row, column=1, value=f"Screen {idx + 1}")
            ws.cell(row=row, column=2, value=inp.product_class)
            ws.cell(row=row, column=3, value=f"{inp.pixel_pitch}mm")
            ws.cell(row=row, column=4, value=f"{inp.width_ft}' × {inp.height_ft}'")
            ws.cell(row=row, column=5, value=f"{inp.width_px} × {inp.height_px}")
            ws.cell(row=row, column=6, value=inp.total_sqft).number_format = "#,##0"
            ws.cell(
                row=row, column=7, value=details.get("hardware", {}).get("unit_cost", 0)
            ).number_format = "$#,##0"
            ws.cell(row=row, column=8, value=hardware_cost).number_format = "$#,##0"

        # Total
        total_row = len(project_data) + 5
        ws.cell(row=total_row, column=1, value="TOTAL LED HARDWARE").font = Font(
            bold=True
        )
        ws.cell(row=total_row, column=8, value=total_hardware).number_format = "$#,##0"
        ws.cell(row=total_row, column=8).font = Font(bold=True)

    def _generate_structural_tab(self, ws, project_data):
        """Structural materials, steel, mounting, engineering"""
        ws.cell(row=1, column=1, value="STRUCTURAL REQUIREMENTS ANALYSIS").font = Font(
            bold=True, size=14
        )

        # Structural materials breakdown
        headers = [
            "Screen",
            "Mounting Type",
            "Structure Cond",
            "Steel Required",
            "Engineering",
            "Structural Mat'ls",
            "Structural Labor",
        ]
        self._set_headers(ws, headers, 3)

        total_structural = 0
        for idx, item in enumerate(project_data):
            row = idx + 4
            inp = item["inputs"]
            details = item["details"]

            structural_materials = (
                details.get("structural", {})
                .get("structural_materials", {})
                .get("raw_cost", 0)
            )
            structural_labor = (
                details.get("structural", {})
                .get("structural_labor", {})
                .get("raw_cost", 0)
            )
            engineering = details.get("engineering", {}).get("raw_cost", 0)
            total_structural += structural_materials + structural_labor + engineering

            ws.cell(row=row, column=1, value=f"Screen {idx + 1}")
            ws.cell(row=row, column=2, value=inp.mounting_type)
            ws.cell(row=row, column=3, value=inp.structure_condition)
            ws.cell(row=row, column=4, value="Calculated").font = Font(italic=True)
            ws.cell(row=row, column=5, value=engineering).number_format = "$#,##0"
            ws.cell(
                row=row, column=6, value=structural_materials
            ).number_format = "$#,##0"
            ws.cell(row=row, column=7, value=structural_labor).number_format = "$#,##0"

        # Installation complexity assessment
        ws.cell(
            row=len(project_data) + 7, column=1, value="INSTALLATION COMPLEXITY FACTORS"
        ).font = Font(bold=True, size=12)
        complexity_factors = [
            "Indoor/Outdoor Environment",
            "Front/Rear Service Access",
            "Crane Access Required",
            "Existing Steel Structure",
            "Power Distance from Panel",
            "Working Hours Restrictions",
            "Weather Protection Required",
            "Permit Complexity",
        ]

        start_row = len(project_data) + 8
        for idx, factor in enumerate(complexity_factors):
            ws.cell(row=start_row + idx, column=1, value=factor)
            ws.cell(row=start_row + idx, column=2, value="Assessed").font = Font(
                italic=True
            )

        # Total
        total_row = len(project_data) + 17
        ws.cell(row=total_row, column=1, value="TOTAL STRUCTURAL").font = Font(
            bold=True
        )
        ws.cell(
            row=total_row, column=7, value=total_structural
        ).number_format = "$#,##0"
        ws.cell(row=total_row, column=7).font = Font(bold=True)

    def _generate_labor_tab(self, ws, project_data):
        """Labor analysis with union rates and crew sizes"""
        ws.cell(
            row=1, column=1, value="LABOR ANALYSIS & CREW REQUIREMENTS"
        ).font = Font(bold=True, size=14)

        headers = [
            "Screen",
            "Labor Type",
            "Crew Size",
            "Duration (Days)",
            "Hourly Rate",
            "Total Hours",
            "Labor Cost",
        ]
        self._set_headers(ws, headers, 3)

        total_labor = 0
        for idx, item in enumerate(project_data):
            row = idx + 4
            inp = item["inputs"]
            details = item["details"]

            # LED installation labor
            led_labor = details.get("led_installation", {}).get("raw_cost", 0)
            electrical_labor = (
                details.get("electrical", {})
                .get("electrical_labor", {})
                .get("raw_cost", 0)
            )
            total_screen_labor = led_labor + electrical_labor
            total_labor += total_screen_labor

            ws.cell(row=row, column=1, value=f"Screen {idx + 1}")
            ws.cell(row=row, column=2, value=inp.labor_type)
            ws.cell(row=row, column=3, value="Calculated").font = Font(italic=True)
            ws.cell(row=row, column=4, value="Calculated").font = Font(italic=True)
            ws.cell(row=row, column=5, value="Union Rate").font = Font(italic=True)
            ws.cell(row=row, column=6, value="Calculated").font = Font(italic=True)
            ws.cell(
                row=row, column=7, value=total_screen_labor
            ).number_format = "$#,##0"

        # Union rate table
        ws.cell(
            row=len(project_data) + 7, column=1, value="UNION RATE REFERENCE"
        ).font = Font(bold=True, size=12)
        union_rates = [
            ("Non-Union", "$45/hour"),
            ("Union - Local", "$75/hour"),
            ("Union - Travel", "$95/hour"),
            ("Prevailing Wage", "$85/hour"),
        ]

        start_row = len(project_data) + 8
        for idx, (union_type, rate) in enumerate(union_rates):
            ws.cell(row=start_row + idx, column=1, value=union_type)
            ws.cell(row=start_row + idx, column=2, value=rate)

        # Total
        total_row = len(project_data) + 14
        ws.cell(row=total_row, column=1, value="TOTAL LABOR").font = Font(bold=True)
        ws.cell(row=total_row, column=7, value=total_labor).number_format = "$#,##0"
        ws.cell(row=total_row, column=7).font = Font(bold=True)

    def _generate_electrical_tab(self, ws, project_data):
        """Electrical systems, power distribution, data infrastructure"""
        ws.cell(row=1, column=1, value="ELECTRICAL SYSTEMS ANALYSIS").font = Font(
            bold=True, size=14
        )

        headers = [
            "Screen",
            "Power Req (kW)",
            "Panel Distance",
            "Electrical Mat'ls",
            "Electrical Labor",
            "Data Infrastructure",
            "Total Electrical",
        ]
        self._set_headers(ws, headers, 3)

        total_electrical = 0
        for idx, item in enumerate(project_data):
            row = idx + 4
            inp = item["inputs"]
            details = item["details"]

            electrical_materials = (
                details.get("electrical", {})
                .get("electrical_materials", {})
                .get("raw_cost", 0)
            )
            electrical_labor = (
                details.get("electrical", {})
                .get("electrical_labor", {})
                .get("raw_cost", 0)
            )
            total_electrical += electrical_materials + electrical_labor

            ws.cell(row=row, column=1, value=f"Screen {idx + 1}")
            ws.cell(row=row, column=2, value="Calculated").font = Font(italic=True)
            ws.cell(
                row=row, column=3, value=inp.power_distance
            ).number_format = "$#,##0"
            ws.cell(
                row=row, column=4, value=electrical_materials
            ).number_format = "$#,##0"
            ws.cell(row=row, column=5, value=electrical_labor).number_format = "$#,##0"
            ws.cell(row=row, column=6, value="Included").font = Font(italic=True)
            ws.cell(
                row=row, column=7, value=electrical_materials + electrical_labor
            ).number_format = "$#,##0"

        # Power requirements table
        ws.cell(
            row=len(project_data) + 7,
            column=1,
            value="POWER REQUIREMENTS BY DISPLAY TYPE",
        ).font = Font(bold=True, size=12)
        power_reqs = [
            ("Indoor 4mm", "0.5 kW/sqft"),
            ("Indoor 6mm", "0.4 kW/sqft"),
            ("Outdoor 10mm", "0.8 kW/sqft"),
            ("Outdoor 16mm", "0.6 kW/sqft"),
        ]

        start_row = len(project_data) + 8
        for idx, (display_type, power_req) in enumerate(power_reqs):
            ws.cell(row=start_row + idx, column=1, value=display_type)
            ws.cell(row=start_row + idx, column=2, value=power_req)

        # Total
        total_row = len(project_data) + 14
        ws.cell(row=total_row, column=1, value="TOTAL ELECTRICAL").font = Font(
            bold=True
        )
        ws.cell(
            row=total_row, column=7, value=total_electrical
        ).number_format = "$#,##0"
        ws.cell(row=total_row, column=7).font = Font(bold=True)

    def _generate_installation_tab(self, ws, project_data):
        """Installation assessment, site conditions, complexity factors"""
        ws.cell(
            row=1, column=1, value="INSTALLATION ASSESSMENT & SITE CONDITIONS"
        ).font = Font(bold=True, size=14)

        # Installation factors table
        ws.cell(row=3, column=1, value="INSTALLATION COMPLEXITY FACTORS").font = Font(
            bold=True, size=12
        )

        installation_factors = [
            ("Environment", "Indoor/Outdoor Assessment"),
            ("Access Method", "Front/Rear Service Access"),
            ("Mounting Type", "Wall/Pole/Header/Floor"),
            ("Structure Condition", "Existing/New Steel"),
            ("Crane Required", "Yes/No - Equipment Access"),
            ("Power Distance", "Close/Medium/Far from Panel"),
            ("Working Hours", "Standard/Restricted/Overtime"),
            ("Weather Protection", "Required/Not Required"),
            ("Permit Complexity", "Simple/Moderate/Complex"),
            ("Site Prep Required", "Minimal/Moderate/Extensive"),
        ]

        for idx, (factor, description) in enumerate(installation_factors):
            row = idx + 4
            ws.cell(row=row, column=1, value=factor).font = Font(bold=True)
            ws.cell(row=row, column=2, value=description)
            ws.cell(row=row, column=3, value="Assessed").font = Font(italic=True)
            ws.cell(row=row, column=4, value="Impact").font = Font(italic=True)

        # Site conditions assessment
        ws.cell(row=15, column=1, value="SITE CONDITION ASSESSMENT").font = Font(
            bold=True, size=12
        )

        site_conditions = [
            "Venue Type (NFL/NBA/NCAA/Other)",
            "Ceiling Height Restrictions",
            "Existing Rigging Points",
            "Electrical Panel Location",
            "Network Infrastructure",
            "Staging Area Available",
            "Security Requirements",
        ]

        for idx, condition in enumerate(site_conditions):
            row = idx + 16
            ws.cell(row=row, column=1, value=condition)
            ws.cell(row=row, column=2, value="Evaluated").font = Font(italic=True)

        # Column widths
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 35
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 15

    def _generate_professional_services_tab(self, ws, project_data):
        """Professional services: PM, engineering, permits, commissioning"""
        ws.cell(row=1, column=1, value="PROFESSIONAL SERVICES BREAKDOWN").font = Font(
            bold=True, size=14
        )

        headers = ["Service Category", "Description", "Duration", "Rate", "Total Cost"]
        self._set_headers(ws, headers, 3)

        total_services = 0

        # Project Management
        pm_cost = sum(
            item["details"].get("project_management", {}).get("raw_cost", 0)
            for item in project_data
        )
        ws.cell(row=4, column=1, value="Project Management")
        ws.cell(row=4, column=2, value="Overall project coordination and management")
        ws.cell(row=4, column=3, value="Project Duration").font = Font(italic=True)
        ws.cell(row=4, column=4, value="% of Total").font = Font(italic=True)
        ws.cell(row=4, column=5, value=pm_cost).number_format = "$#,##0"
        total_services += pm_cost

        # Engineering
        engineering_cost = sum(
            item["details"].get("engineering", {}).get("raw_cost", 0)
            for item in project_data
        )
        ws.cell(row=5, column=1, value="Engineering Services")
        ws.cell(row=5, column=2, value="Structural and electrical engineering")
        ws.cell(row=5, column=3, value="Design Phase").font = Font(italic=True)
        ws.cell(row=5, column=4, value="Fixed Fee").font = Font(italic=True)
        ws.cell(row=5, column=5, value=engineering_cost).number_format = "$#,##0"
        total_services += engineering_cost

        # Permits
        permits_cost = sum(
            item["details"].get("permits", {}).get("raw_cost", 0)
            for item in project_data
        )
        ws.cell(row=6, column=1, value="Permits & Approvals")
        ws.cell(row=6, column=2, value="Building permits and inspections")
        ws.cell(row=6, column=3, value="Permit Duration").font = Font(italic=True)
        ws.cell(row=6, column=4, value="Actual Cost").font = Font(italic=True)
        ws.cell(row=6, column=5, value=permits_cost).number_format = "$#,##0"
        total_services += permits_cost

        # Submittals
        submittals_cost = sum(
            item["details"].get("submittals", {}).get("raw_cost", 0)
            for item in project_data
        )
        ws.cell(row=7, column=1, value="Submittals")
        ws.cell(row=7, column=2, value="Technical submittals and drawings")
        ws.cell(row=7, column=3, value="Submittal Phase").font = Font(italic=True)
        ws.cell(row=7, column=4, value="Fixed Fee").font = Font(italic=True)
        ws.cell(row=7, column=5, value=submittals_cost).number_format = "$#,##0"
        total_services += submittals_cost

        # Final Commissioning
        commissioning_cost = sum(
            item["details"].get("final_commissioning", {}).get("raw_cost", 0)
            for item in project_data
        )
        ws.cell(row=8, column=1, value="Final Commissioning")
        ws.cell(row=8, column=2, value="System testing and commissioning")
        ws.cell(row=8, column=3, value="Commissioning Period").font = Font(italic=True)
        ws.cell(row=8, column=4, value="Fixed Fee").font = Font(italic=True)
        ws.cell(row=8, column=5, value=commissioning_cost).number_format = "$#,##0"
        total_services += commissioning_cost

        # Total
        total_row = 10
        ws.cell(
            row=total_row, column=1, value="TOTAL PROFESSIONAL SERVICES"
        ).font = Font(bold=True)
        ws.cell(row=total_row, column=5, value=total_services).number_format = "$#,##0"
        ws.cell(row=total_row, column=5).font = Font(bold=True)

    def _generate_screen_detail_tab(self, ws, item, screen_num):
        """Detailed cost breakdown for individual screen (existing enhanced)"""
        # Title
        inp = item["inputs"]
        ws.cell(
            row=1,
            column=1,
            value=f"Screen {screen_num}: {inp.product_class} - Detailed Analysis",
        ).font = Font(bold=True, size=14)

        # Basic specs
        ws.cell(row=3, column=1, value="Screen Specifications:").font = Font(bold=True)
        specs = [
            ("Product Class", inp.product_class),
            ("Pixel Pitch", f"{inp.pixel_pitch}mm"),
            ("Dimensions", f"{inp.width_ft}' × {inp.height_ft}'"),
            ("Resolution", f"{inp.width_px} × {inp.height_px} pixels"),
            ("Total Square Footage", f"{inp.total_sqft:,.0f} sq ft"),
            ("Indoor/Outdoor", "Indoor" if inp.indoor else "Outdoor"),
            ("Mounting Type", inp.mounting_type),
            ("Structure Condition", inp.structure_condition),
        ]

        for idx, (label, value) in enumerate(specs):
            ws.cell(row=idx + 4, column=1, value=label).font = Font(bold=True)
            ws.cell(row=idx + 4, column=2, value=value)

        # Detailed cost breakdown
        start_row = 13
        ws.cell(row=start_row, column=1, value="Detailed Cost Breakdown:").font = Font(
            bold=True, size=12
        )

        headers = [
            "Category",
            "Description",
            "Raw Cost",
            "Calculation Logic",
            "Sell Price",
        ]
        self._set_headers(ws, headers, start_row + 1)

        # Generate detailed breakdown using existing logic
        details = item["details"]
        breakdown = item["cost_breakdown"]

        mapping = [
            ("1. Hardware", "hardware", None),
            ("2. Structural Materials", "structural", "structural_materials"),
            ("3. Structural Labor", "structural", "structural_labor"),
            ("4. LED Installation", "led_installation", None),
            ("5. Electrical Materials", "electrical", "electrical_materials"),
            ("6. Electrical Labor", "electrical", "electrical_labor"),
            ("7. CMS Equipment", "cms", "cms_equipment"),
            ("8. CMS Installation", "cms", "cms_installation"),
            ("9. CMS Commissioning", "cms", "cms_commissioning"),
            ("10. Project Management", "project_management", None),
            ("11. General Conditions", "general_conditions", None),
            ("12. Travel & Expenses", "travel", None),
            ("13. Submittals", "submittals", None),
            ("14. Engineering", "engineering", None),
            ("15. Permits", "permits", None),
            ("16. Final Commissioning", "final_commissioning", None),
        ]

        row = start_row + 2
        for label, detail_key, sub_key in mapping:
            sell_price = breakdown.get(label, 0)
            raw_obj = {}
            if detail_key in details:
                parent = details[detail_key]
                if sub_key:
                    raw_obj = parent.get(sub_key, {})
                else:
                    raw_obj = parent

            description = raw_obj.get("description", "")
            raw_cost = raw_obj.get("raw_cost", 0)
            logic = raw_obj.get("calculation", "")

            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=description)
            ws.cell(row=row, column=3, value=raw_cost).number_format = "$#,##0"
            ws.cell(row=row, column=4, value=logic)
            ws.cell(row=row, column=5, value=sell_price).number_format = "$#,##0"
            row += 1

        # Bond & Contingency
        ws.cell(row=row, column=1, value="17. Bond")
        ws.cell(
            row=row, column=5, value=breakdown.get("17. Bond", 0)
        ).number_format = "$#,##0"
        row += 1

        ws.cell(row=row, column=1, value="18. Contingency")
        ws.cell(
            row=row,
            column=4,
            value=f"{item['pricing'].get('contingency_pct', 0) * 100}% of Subtotal",
        )
        ws.cell(
            row=row, column=5, value=breakdown.get("18. Contingency", 0)
        ).number_format = "$#,##0"
        row += 1

        # Total
        row += 1
        ws.cell(row=row, column=1, value="TOTAL SELL PRICE").font = Font(
            bold=True, size=12
        )
        ws.cell(
            row=row, column=5, value=item["summary"]["final_sell_price"]
        ).number_format = "$#,##0"
        ws.cell(row=row, column=5).font = Font(bold=True, size=14)

    def _set_headers(self, ws, headers, row):
        """Helper function to set table headers"""
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(
                start_color="003D82", end_color="003D82", fill_type="solid"
            )
            cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[get_column_letter(col)].width = 20

        # Wider columns for descriptions
        if len(headers) > 1:
            ws.column_dimensions["B"].width = 35
        if len(headers) > 3:
            ws.column_dimensions["D"].width = 45
