"""Generate a dummy Master Excel used as placeholder until real Master Excel is provided.
Creates `samples/master_excel_dummy.xlsx` with basic tabs and precomputed values + formula strings for human review.
"""
from openpyxl import Workbook
from pathlib import Path

OUT_PATH = Path("samples")
OUT_PATH.mkdir(exist_ok=True)
OUT_FILE = OUT_PATH / "master_excel_dummy.xlsx"

wb = Workbook()

# Hardware sheet
ws = wb.active
ws.title = "Hardware"
ws.append(["screen_id","product_class","width_ft","height_ft","pixel_pitch","base_rate","sqft","hardware_cost","hardware_formula"])
rows = [
    [1, "Ribbon", 40, 6, 10, 1200],
    [2, "Scoreboard", 30, 16, 6, 2400],
]
for r in rows:
    screen_id, pc, w, h, pitch, base = r
    sqft = w * h
    hw_cost = sqft * base
    formula = f"=C{ws.max_row+1}*D{ws.max_row+1}*F{ws.max_row+1}"
    ws.append([screen_id, pc, w, h, pitch, base, sqft, hw_cost, formula])

# Structural sheet
ws2 = wb.create_sheet(title="Structural")
ws2.append(["screen_id","structural_pct","structural_cost_formula","structural_cost"])
for idx, r in enumerate(rows, start=1):
    structural_pct = 0.20
    # formula reference to Hardware.hardware_cost
    formula = f"=Hardware!H{idx+1}*{structural_pct}"
    # Compute structural cost using the same numbers
    structural_cost = (r[2] * r[3] * r[5]) * structural_pct
    ws2.append([idx, structural_pct, formula, structural_cost])

# Labor sheet
ws3 = wb.create_sheet(title="Labor")
ws3.append(["screen_id","labor_pct","labor_cost_formula","labor_cost"])
for idx, r in enumerate(rows, start=1):
    labor_pct = 0.15
    formula = f"=(Hardware!H{idx+1} + Structural!D{idx+1})*{labor_pct}"
    labor_cost = ((r[2] * r[3] * r[5]) + ((r[2] * r[3] * r[5]) * 0.20)) * labor_pct
    ws3.append([idx, labor_pct, formula, labor_cost])

wb.save(OUT_FILE)
print(f"Dummy master excel created: {OUT_FILE}")
