"""Generate a more realistic Master Excel similar to what Natalia might have.
Creates `samples/master_excel_realistic_dummy.xlsx` with multiple tabs and formula strings for human review.
"""
from openpyxl import Workbook
from pathlib import Path
from datetime import date

OUT_PATH = Path("samples")
OUT_PATH.mkdir(exist_ok=True)
OUT_FILE = OUT_PATH / "master_excel_realistic_dummy.xlsx"

wb = Workbook()

# Metadata sheet
meta = wb.active
meta.title = "Metadata"
meta.append(["Master Sheet", "ANC Realistic Dummy"])
meta.append(["Version", "1.0"])
meta.append(["Generated", date.today().isoformat()])

# Hardware sheet
hw = wb.create_sheet(title="Hardware")
hw.append(["screen_id","product_code","product_class","width_ft","height_ft","pixel_pitch","base_rate_per_sqft","sqft","unit_qty","hardware_cost","hardware_formula","notes"])
rows = [
    [1, "RIB-10-40x6", "Ribbon", 40, 6, 10, 120.0, None, 1, None, None, "Indoor ribbon"],
    [2, "SCB-6-30x16", "Scoreboard", 30, 16, 6, 240.0, None, 1, None, None, "Outdoor scoreboard (weather-rated)"]
]
for r in rows:
    screen_id, code, pc, w, h, pitch, rate, _, unit = r[0], r[1], r[2], r[3], r[4], r[5], r[6], r[7], r[8]
    sqft = w * h
    hw_cost = sqft * rate * unit
    formula = f"=D{hw.max_row+1}*E{hw.max_row+1}*G{hw.max_row+1}*I{hw.max_row+1}"
    hw.append([screen_id, code, pc, w, h, pitch, rate, sqft, unit, hw_cost, formula, r[11] if len(r) > 11 else ""])

# Structural sheet
st = wb.create_sheet(title="Structural")
st.append(["screen_id","structural_pct","structural_cost_formula","structural_cost","notes"])
for idx, r in enumerate(rows, start=1):
    structural_pct = 0.20
    formula = f"=Hardware!J{idx+1}*{structural_pct}"
    structural_cost = (r[3] * r[4] * r[6] * r[8]) * structural_pct
    st.append([idx, structural_pct, formula, structural_cost, "Includes steel and brackets"])

# Labor sheet
lab = wb.create_sheet(title="Labor")
lab.append(["screen_id","labor_pct","labor_cost_formula","labor_cost","crew_size","labor_hours","notes"])
for idx, r in enumerate(rows, start=1):
    labor_pct = 0.15
    formula = f"=(Hardware!J{idx+1} + Structural!D{idx+1})*{labor_pct}"
    labor_cost = ((r[3] * r[4] * r[6] * r[8]) + ((r[3] * r[4] * r[6] * r[8]) * 0.20)) * labor_pct
    crew = 3 if r[2] == "Ribbon" else 4
    hours = crew * 8
    lab.append([idx, labor_pct, formula, labor_cost, crew, hours, "Includes installation and setup"])

# Shipping & Handling
ship = wb.create_sheet(title="Shipping")
ship.append(["screen_id","shipping_pct","shipping_cost_formula","shipping_cost","notes"])
for idx, r in enumerate(rows, start=1):
    shipping_pct = 0.02
    formula = f"=Hardware!J{idx+1}*{shipping_pct}"
    shipping_cost = (r[3] * r[4] * r[6] * r[8]) * shipping_pct
    ship.append([idx, shipping_pct, formula, shipping_cost, "Freight and crating"])

# Misc Expenses
misc = wb.create_sheet(title="Misc")
misc.append(["screen_id","misc_fixed","misc_cost_formula","misc_cost","notes"])
for idx, r in enumerate(rows, start=1):
    misc_fixed = 500
    formula = f"={misc_fixed}"
    misc.append([idx, misc_fixed, formula, misc_fixed, "Permits, misc consumables"])

# Contingency/Margins
finance = wb.create_sheet(title="Finance")
finance.append(["item","value","formula","notes"])
# We'll set contingency, bond, margin
contingency_pct = 0.05
bond_pct = 0.02
margin_pct = 0.20
finance.append(["contingency_pct", contingency_pct, f"={contingency_pct}", "Project risk contingency"])
finance.append(["bond_pct", bond_pct, f"={bond_pct}", "Performance bond"])
finance.append(["margin_pct", margin_pct, f"={margin_pct}", "Target gross margin"])

# Summary per screen
sumsh = wb.create_sheet(title="Summary")
sumsh.append(["screen_id","hardware","structural","labor","shipping","misc","subtotal","contingency","bond","margin","total_formula","total_value","notes"])
for idx, r in enumerate(rows, start=1):
    # references
    h_ref = f"Hardware!J{idx+1}"
    s_ref = f"Structural!D{idx+1}"
    l_ref = f"Labor!D{idx+1}"
    sh_ref = f"Shipping!D{idx+1}"
    m_ref = f"Misc!C{idx+1}"  # misc cost
    subtotal_formula = f"={h_ref}+{s_ref}+{l_ref}+{sh_ref}+{m_ref}"
    # contingency, bond, margin applied on subtotal
    total_formula = f"=({subtotal_formula})*(1+Finance!B1)*(1+Finance!B2)*(1+Finance!B3)"
    # computed values
    h_val = (r[3] * r[4] * r[6] * r[8])
    s_val = h_val * 0.20
    l_val = (h_val + s_val) * 0.15
    sh_val = h_val * 0.02
    m_val = 500
    subtotal_val = h_val + s_val + l_val + sh_val + m_val
    total_val = subtotal_val * (1 + contingency_pct) * (1 + bond_pct) * (1 + margin_pct)
    sumsh.append([idx, h_val, s_val, l_val, sh_val, m_val, subtotal_val, f"=G{sumsh.max_row+1}*Finance!B1", f"=G{sumsh.max_row+1}*Finance!B2", f"=G{sumsh.max_row+1}*Finance!B3", total_formula, total_val, "Per-screen totals"])

# Grand totals
sumsh.append([])
sumsh.append(["GRAND_TOTAL", "", "", "", "", "", f"=SUM(G2:G{1+len(rows)})", "=G{len(rows)+2}*Finance!B1", "=G{len(rows)+2}*Finance!B2", "=G{len(rows)+2}*Finance!B3", f"=G{len(rows)+2}*(1+Finance!B1)*(1+Finance!B2)*(1+Finance!B3)", "=SUM(L2:L{1+len(rows)})", "Project grand totals"])

wb.save(OUT_FILE)
print(f"Realistic dummy master excel created: {OUT_FILE}")
